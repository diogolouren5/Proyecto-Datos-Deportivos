import garminconnect
import openpyxl
import schedule
from datetime import time, datetime, timedelta

# Configura tus credenciales de Garmin Connect
email = "diogolouren5@gmail.com"
password = "Sebastiao552"

def seconds_to_hms(seconds):
    hours = seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = seconds % 60
    seconds=round(seconds)
    return f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}"

try:
    
    # Autenticar con Garmin Connect
    print("Iniciando sesión en Garmin Connect...")
    garmin = garminconnect.Garmin(email, password)
    garmin.login()
    garmin.display_name
    print("¡Inicio de sesión exitoso en Garmin Connect!\n")

    # Cargar plantilla de Excel
    file_path = "C:/Users/Diogo/Desktop/IronMan 2025/GarminRawData.xlsx"
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Obtener actividades recientes
    L_A = garmin.get_activities(0, 1)  # Obtiene la actividad más recientes
    
    for datos in L_A:

        L_A_ID=datos.get("activityId")
        print(f"ID de la última actividad: {L_A_ID}")

        L_A_Type=datos['activityType']['typeKey']
        print(f"Tipo de Actividad: {L_A_Type}")

        L_A_Name=datos.get("activityName")
        print(f"Nombre de la actividad: {L_A_Name}")

        L_A_Start_Time=datos.get("startTimeLocal")
        print(f"Start Time de la actividad: {L_A_Start_Time}")

    print(f"\n¡La actividad ha sido reconocida correctamente!\n")


    #Asociar la fecha de la actividad a la fila de ese dia en excel
    # Convertir fecha a formato datetime
    activity_date = datetime.strptime(L_A_Start_Time, "%Y-%m-%d %H:%M:%S")

    # Formatear la fecha como dd/mm/aaaa para Excel
    formatted_date = activity_date.strftime('%d/%m/%Y')
    print(f"Fecha Formateada: {formatted_date}")

    # Encontrar la target row con dias ordinales del año
    target_row = activity_date.timetuple().tm_yday+21
    print(f"Fila objetivo con la fecha buscada: {target_row}")



    if L_A_Type in ['swimming',"lap_swimming","pool_swimming","open_water_swimming"]:
        
        for datos in L_A:
            L_A_distance = datos.get("distance")
            L_A_duration = seconds_to_hms(datos.get("duration"))
            L_A_movingDuration = seconds_to_hms(datos.get("movingDuration"))
            L_A_averageSpeed = round(datos.get("averageSpeed")*(36/10),2)
            L_A_maxSpeed = round(datos.get("maxSpeed")*(36/10),2)
            L_A_averageHR = datos.get("averageHR")
            L_A_maxHR = datos.get("maxHR")
            L_A_averageSwimCadenceInStrokesPerMinute = datos.get("averageSwimCadenceInStrokesPerMinute")
            L_A_averageSwolf = datos.get("averageSwolf")
            L_A_activeLengths = datos.get("activeLengths")
            L_A_strokes = datos.get("strokes")
            L_A_avgStrokes = round(datos.get("avgStrokes"),2)



        ### Meter los datos en su celda correspondiente al parametro y fecha

        ws.cell(row=target_row, column=7).value = L_A_distance           
        ws.cell(row=target_row, column=8).value = L_A_duration          
        ws.cell(row=target_row, column=9).value = L_A_movingDuration     
        ws.cell(row=target_row, column=10).value = L_A_averageSpeed        
        ws.cell(row=target_row, column=11).value = L_A_maxSpeed    
        ws.cell(row=target_row, column=12).value = L_A_averageHR           
        ws.cell(row=target_row, column=13).value =  L_A_maxHR
        ws.cell(row=target_row, column=14).value =  L_A_averageSwimCadenceInStrokesPerMinute
        ws.cell(row=target_row, column=15).value =  L_A_averageSwolf
        ws.cell(row=target_row, column=16).value =  L_A_activeLengths
        ws.cell(row=target_row, column=17).value =  L_A_strokes
        ws.cell(row=target_row, column=18).value =  L_A_avgStrokes

        
    elif L_A_Type in["running","treadmill_running" , "trail_running","track_running","coss_country", "road_running" ]:
        for datos in L_A:
            L_A_distance=round(datos.get("distance")/1000,2)
            L_A_duration=seconds_to_hms(datos.get("duration"))
            L_A_movingDuration=seconds_to_hms(datos.get("movingDuration"))
            L_A_averageSpeed=round(datos.get("averageSpeed")*(36/10),2)
            L_A_maxSpeed=round(datos.get("maxSpeed")*(36/10),2)
            L_A_averageHR=datos.get("averageHR")
            L_A_maxHR=datos.get("maxHR")
            L_A_averageRunningCadenceInStepsPerMinute=round(datos.get("averageRunningCadenceInStepsPerMinute"),2)
            L_A_maxRunningCadenceInStepsPerMinute=round(datos.get("maxRunningCadenceInStepsPerMinute"),2)
            L_A_steps=datos.get("steps")
            L_A_avgPower=round(datos.get("avgPower"),2)
            L_A_maxPower=round(datos.get("maxPower"),2)
            L_A_normPower=round(datos.get("normPower"),2)
            L_A_avgVerticalOscillation=round(datos.get("avgVerticalOscillation"),2)
            L_A_avgGroundContactTime=round(datos.get("avgGroundContactTime"),2)
            L_A_avgStrideLength=round(datos.get("avgStrideLength"),2)
            L_A_avgRespirationRate=round(datos.get("avgRespirationRate"),2)
            L_A_maxRespirationRate=round(datos.get("maxRespirationRate"),2)
            L_A_trainingEffectLabel=datos.get("trainingEffectLabel")
            
       
        ### Meter los datos en su celda correspondiente al parametro y fecha   
        ws.cell(row=target_row, column=35).value = L_A_distance
        ws.cell(row=target_row, column=36).value = L_A_duration
        ws.cell(row=target_row, column=37).value = L_A_movingDuration
        ws.cell(row=target_row, column=38).value = L_A_averageSpeed
        ws.cell(row=target_row, column=39).value = L_A_maxSpeed
        ws.cell(row=target_row, column=40).value = L_A_averageHR
        ws.cell(row=target_row, column=41).value = L_A_maxHR
        ws.cell(row=target_row, column=42).value = L_A_averageRunningCadenceInStepsPerMinute
        ws.cell(row=target_row, column=43).value = L_A_maxRunningCadenceInStepsPerMinute
        ws.cell(row=target_row, column=44).value = L_A_steps
        ws.cell(row=target_row, column=45).value = L_A_avgPower
        ws.cell(row=target_row, column=46).value = L_A_maxPower
        ws.cell(row=target_row, column=47).value = L_A_normPower
        ws.cell(row=target_row, column=48).value =L_A_avgVerticalOscillation
        ws.cell(row=target_row, column=49).value = L_A_avgGroundContactTime
        ws.cell(row=target_row, column=50).value = L_A_avgStrideLength
        ws.cell(row=target_row, column=51).value = L_A_avgRespirationRate
        ws.cell(row=target_row, column=52).value = L_A_maxRespirationRate
        ws.cell(row=target_row, column=53).value = L_A_trainingEffectLabel

        
    elif L_A_Type in ["cycling","road_biking" ,"mountain_biking" ,"cyclocross" ,"virtual_ride" ,"indoor_cycling", "track_cycling" , "stationary_cycling"]:
       
        for datos in L_A:
            L_A_distance=round(datos.get("distance")/1000,2)
            L_A_duration=seconds_to_hms(datos.get("duration"))
            L_A_movingDuration=seconds_to_hms(datos.get("movingDuration"))
            L_A_elevationGain=round(datos.get("elevationGain"),2)
            L_A_averageSpeed=round(datos.get("averageSpeed")*(36/10),2)
            L_A_maxSpeed=round(datos.get("maxSpeed")*(36/10),2)
            L_A_averageHR=datos.get("averageHR")
            L_A_maxHR=datos.get("maxHR")
            L_A_maxFtp=round(datos.get("maxFtp"),2)
            L_A_maxTemperature=datos.get("maxTemperature")
            L_A_maxRespirationRate=round(datos.get("maxRespirationRate"),2)
            L_A_avgRespirationRate=round(datos.get("avgRespirationRate"),2)
            """L_A_avgCadence=round(datos.get("avgCadence",0))
            L_A_avgPower=round(datos.get("avgPower"),2)
            L_A_maxPower=round(datos.get("maxPower"),2)
            L_A_normPower=round(datos.get("normPower"),2)"""
           

        ### Meter los datos en su celda correspondiente al parametro y fecha
        ws.cell(row=target_row, column=19).value = L_A_distance      
        ws.cell(row=target_row, column=20).value = L_A_duration
        ws.cell(row=target_row, column=21).value = L_A_movingDuration
        ws.cell(row=target_row, column=22).value = L_A_elevationGain
        ws.cell(row=target_row, column=23).value = L_A_averageSpeed
        ws.cell(row=target_row, column=24).value = L_A_maxSpeed
        ws.cell(row=target_row, column=25).value = L_A_averageHR
        ws.cell(row=target_row, column=26).value = L_A_maxHR
        ws.cell(row=target_row, column=27).value = L_A_maxFtp
        ws.cell(row=target_row, column=28).value = L_A_maxTemperature
        ws.cell(row=target_row, column=29).value = L_A_maxRespirationRate
        ws.cell(row=target_row, column=30).value = L_A_avgRespirationRate
        """ws.cell(row=target_row, column=31).value = L_A_avgCadence
        ws.cell(row=target_row, column=32).value = L_A_avgPower
        ws.cell(row=target_row, column=33).value =L_A_maxPower
        ws.cell(row=target_row, column=34).value = L_A_normPower"""



    # Actualizar la celda B2 con la fecha y hora actual
    ws['B2'] = datetime.now().strftime('%d/%m/%Y %H:%M:%S')

    # Guardar el archivo de Excel
    wb.save(file_path)
    print("\nArchivo Excel actualizado correctamente.")

except Exception as e:
    print("\nError al ejecutar el código:")
    print(str(e))
